import time

import openpyxl
from selenium.webdriver.chrome import webdriver
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

file_path="/Users/ntalesha/Downloads/download.xlsx"

def update_excel_data(filepath, columnName, fruitName, price):
    book = openpyxl.load_workbook(filepath)
    sheet = book.active
    dic = {}
    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == columnName:
            dic["col"] = i

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == fruitName:
                dic["row"] = i

    sheet.cell(row=dic["row"], column=dic["col"]).value = price
    book.save(file_path)

driver=webdriver.Chrome()
driver.implicitly_wait(10)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.find_element(By.CSS_SELECTOR,"#downloadButton").click()

#edit excel



update_excel_data(file_path,"price","Apple","900")




file_input = driver.find_element(By.CSS_SELECTOR,"#fileinput")
file_input.send_keys(file_path)

WebDriverWait(driver,10).until(expected_conditions.visibility_of_element_located((By.XPATH,"//div[text()='Updated Excel Data Successfully.']")))
print(driver.find_element(By.XPATH,"//div[text()='Updated Excel Data Successfully.']").text)

priceColID = driver.find_element(By.XPATH,"//*[text()='Price']").get_attribute("data-column-id")
fruit_name = "Apple"
priceval = driver.find_element(By.XPATH,"//*[text()='"+fruit_name+"']/../..//div[@id='cell-"+priceColID+"-undefined']/div").text
print(priceval)


time.sleep(7)





